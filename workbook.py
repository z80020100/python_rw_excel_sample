
import string

import numpy as np
from openpyxl import load_workbook


# Error messages
CELL_NAME_ERROR = "cell name is not valid."
ROW_NAME_ERROR = "row name is not valid."
RANGE_NAME_ERROR = "range name is not valid."
UNSUPPORTED_FILE_EXT_ERROR = "unsupported file extension."


# File extensions
FILE_EXT_EXCEL_WORKBOOK = ".xlsx"
FILE_EXT_EXCEL_MACRO_WORKBOOK = ".xlsm"


# Index types
INDEX_TYPE_PANDAS = 0
INDEX_TYPE_OPENPYXL = 1


class Workbook:

    def __init__(self, filename, writable=False, save_formula_to_value=False):
        self.filename = filename
        self.extension = filename[filename.rfind("."):].lower()
        self.read_only = not writable
        self.data_only = save_formula_to_value or self.read_only
        if self.extension == FILE_EXT_EXCEL_MACRO_WORKBOOK:
            self.keep_vba = True
        elif self.extension == FILE_EXT_EXCEL_WORKBOOK:
            self.keep_vba = False
        else:
            raise ValueError(UNSUPPORTED_FILE_EXT_ERROR)
        self.workbook = load_workbook(
            filename=self.filename,
            read_only=self.read_only,
            keep_vba=self.keep_vba,
            data_only=self.data_only)

    def get_sheets(self):
        return self.workbook.sheetnames

    def get_cell(self, sheet_name, cell_name):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        row_index, column_index = Workbook.convert_cell_name_to_index(
            cell_name, INDEX_TYPE_OPENPYXL)
        return sheet.cell(row=row_index, column=column_index).value

    def get_range(self, sheet_name, range_name):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        start_row_index, end_row_index, start_column_index, end_column_index = Workbook.convert_range_name_index(
            range_name, INDEX_TYPE_OPENPYXL)
        rows = sheet.iter_rows(min_row=start_row_index,
                               max_row=end_row_index,
                               min_col=start_column_index,
                               max_col=end_column_index)
        return np.array([[cell.value for cell in row] for row in rows])

    def set_cell(self, sheet_name, cell_name, value):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        row_index, column_index = Workbook.convert_cell_name_to_index(
            cell_name, INDEX_TYPE_OPENPYXL)
        sheet.cell(row=row_index, column=column_index).value = value

    def set_range(self, sheet_name, range_name, value):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        start_row_index, end_row_index, start_column_index, end_column_index = Workbook.convert_range_name_index(
            range_name, INDEX_TYPE_OPENPYXL)
        rows = sheet.iter_rows(min_row=start_row_index,
                               max_row=end_row_index,
                               min_col=start_column_index,
                               max_col=end_column_index)
        for row in rows:
            for cell in row:
                cell.value = value

    def save(self):
        self.workbook.save(self.filename)

    # Use Excel range format and index is for Pandas
    @staticmethod
    def convert_range_name_index(range_name, index_type=INDEX_TYPE_PANDAS):
        cutting_index = 0
        for i, c in enumerate(range_name):
            if c == ":":
                cutting_index = i
                break
        if cutting_index == 0:
            raise ValueError(RANGE_NAME_ERROR)
        # Split range name to start cell and end cell
        start_cell_name = range_name[:cutting_index]
        end_cell_name = range_name[cutting_index + 1:]
        start_row_index, start_column_index = Workbook.convert_cell_name_to_index(
            start_cell_name, index_type)
        end_row_index, end_column_index = Workbook.convert_cell_name_to_index(
            end_cell_name, index_type)
        return (start_row_index, end_row_index, start_column_index,
                end_column_index)

    # Use Excel cell format and index supports Pandas and OpenPyXL
    @staticmethod
    def convert_cell_name_to_index(cell_name, index_type=INDEX_TYPE_PANDAS):
        upper_cell_name = cell_name.upper()
        cutting_index = 0
        for i, c in enumerate(upper_cell_name):
            if c not in string.ascii_letters:
                cutting_index = i
                break
        if cutting_index == 0:
            raise ValueError(CELL_NAME_ERROR)
        # Split cell name to column and row
        column = upper_cell_name[:cutting_index]
        row = upper_cell_name[cutting_index:]
        # Convert column name to index
        column_index = 0
        for i in range(len(column)):
            column_index += ((ord(column[i]) - ord("A")) + 26 * i)
        # Convert row name to index
        row_index = 0
        try:
            row_index = int(row) - 1
        except ValueError:
            raise ValueError(ROW_NAME_ERROR)
        if row_index < 0:
            raise ValueError(ROW_NAME_ERROR)
        if index_type == INDEX_TYPE_OPENPYXL:
            return row_index + 1, column_index + 1
        else:
            return row_index, column_index
